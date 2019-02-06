# Using Excel with Python
# If a file downloaded from the Internet is not working, use EXCEL to convert the file to the newest xlsx format

import openpyxl

book = openpyxl.load_workbook("stat_104102.xlsx")

# method 1
print(book.get_sheet_names())  
print(book.get_sheet_by_name("Sheet0"))

# method 2
sheet = book.worksheets[0]  
for row in sheet.rows: # In Openpyxl, methods are usually iterable; so use for statement to access items
    for data in row:
        print(data.value, end=" ")
    print()

### Writing to Excel: Refer to the textbook

workbook = openpyxl.Workbook() # Class instance, so it is Workbook not workbook to create a new workbook
# workbook = openpyxl.load_workbook("filename") to open an exising file 
sheet = workbook.active
# writing 
sheet["A1"] = 100
sheet["B1"] = 200
sheet.merge_cells("A1:C1")
sheet["A1"].font = openpyxl.styles.Font(size=20, color="FF0000")

# saving
workbook.save("newFile.xlsx")
